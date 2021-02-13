import argparse
from cv2 import imread, imshow, resize, waitKey
from openpyxl.workbook import Workbook
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import colors
from openpyxl.cell import Cell
import click
from gooey import Gooey, GooeyParser

def center_crop(img, dim):
    """Returns center cropped image
    Args:
    img: image to be center cropped
    dim: dimensions (width, height) to be cropped
    """

# https://gist.github.com/Nannigalaxy/35dd1d0722f29672e68b700bc5d44767

    width, height = img.shape[1], img.shape[0]

    # process crop width and height for max available dimension
    crop_width = dim[0] if dim[0] < img.shape[1] else img.shape[1]
    crop_height = dim[1] if dim[1] < img.shape[0] else img.shape[0]
    mid_x, mid_y = int(width/2), int(height/2)
    cw2, ch2 = int(crop_width/2), int(crop_height/2)
    crop_img = img[mid_y-ch2:mid_y+ch2, mid_x-cw2:mid_x+cw2]
    return crop_img


def RGB_to_HEX(R, G, B):
    """Converts RGB to Hex with no alpha"""

    # slightly modified from https://stackoverflow.com/a/3380739

    return '%02x%02x%02x' % (R, G, B)


def RGB_to_fill(R, G, B):
    """Converts RGB values to a PatternFill"""

    # based on https://stackoverflow.com/q/30484220

    return PatternFill(
        start_color=RGB_to_HEX(R, G, B),
        end_color=RGB_to_HEX(R, G, B),
        fill_type='solid'
    )


def num_to_col(n):
    """Converts a number to a Excel column name"""

    # based on https://stackoverflow.com/a/23862195
    string = ""
    while n > 0:
        n, remainder = divmod(n - 1, 26)
        string = chr(65 + remainder) + string
    return string


def image_to_Excel(input_image_path: str, output_Excel_path: str, x_dim: int, y_dim: int):
    """Takes an image and makes it into an Excel file, where one pixel in the image is one cell."""
    x = x_dim
    y = y_dim
    path = input_image_path

    image = imread(path)
    # uncomment to show image prior to importing it to Excel. For debugging
    # imshow('Window Title', image); waitKey()

    # print info about image
    # this is for debugging
    img_info = image.shape
    print("--- Input Image Info ---")
    print("Image height :", img_info[0])
    print("Image Width :", img_info[1])
    print("Image channels :", img_info[2])

    image = center_crop(image, (x, y))
    img_info = image.shape

    print("\n--- Cropped Image Info ---")
    print("Image height :", img_info[0])
    print("Image Width :", img_info[1])
    print("Image channels :", img_info[2])

    # create an Excel workbook + worksheet
    wb = Workbook()
    ws = wb.active

    # the template for the bar is the default, but it is modifed to use the pipe
    # symbol instead of brackets for the progress bar, and use a block character
    # for progress

    counter = 0
    for a in range(1, x):
        counter += 1
        print(str(counter) + "/" + str(x))
        for b in range(1, y):
            
            
            # store the current pixel in a variable
            pixel = image[a][b]

            # set the fill of the cell to the RGB value of the corresponding
            # pixel in the input image
            # we need to flip R and B for some reason
            ws[num_to_col(b) + str(a)
                ].fill = RGB_to_fill(pixel[2], pixel[1], pixel[0])

            # and make the cells square by setting the column width to 2.54
            # which is the same as the default column height
            ws.column_dimensions[num_to_col(b)].width = 2.54

            

    # save the Excel file
    wb.save("image.xlsx")

    # and open it in Excel
    import os
    os.system(
        r""""c:\Program Files\Microsoft Office\root\Office16\excel.exe" """ + output_Excel_path)

#image_to_Excel(r"c:\Users\Arjun\Documents\Image20210213105453.jpg", r"c:\users\arjun\image.xlsx", 250, 250)

@Gooey(progress_regex=r"^(\d+)/(\d+)$", progress_expr="x[0] / x[1] * 100")
def setupCLI():
    parser = GooeyParser()
    parser.add_argument("imagePath", help=r"The absolute path to the input image", widget='FileChooser')
    parser.add_argument(
        "outputPath", help=r"The absolute path to the resulting Excel file. File will be overwritten if present.", widget='FileChooser')
    parser.add_argument(
        "xDim", type=int, help=r"The desired width of the output, in pixels. Values above 300 produce undefined results.")
    parser.add_argument(
        "yDim", type=int, help=r"The desired height of the output, in pixels. Values above 300 produce undefined results.")
    args = parser.parse_args()

    image_to_Excel(args.imagePath, args.outputPath, args.xDim, args.yDim)
setupCLI()